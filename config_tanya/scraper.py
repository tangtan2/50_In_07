from bs4 import BeautifulSoup, Comment
import config_tanya.scraper1_header
import urllib.request
import shutil
import openpyxl as op
import os
from config_tanya.config import ScraperConfig
from config_tanya.header import connect_db, disconnect_db, get_team_players, import_shortnames, import_longnames


def scrape1(config, url1, short_name):

    # Find overall info for team
    teampage = urllib.request.urlopen(url1)
    teamsoup = BeautifulSoup(teampage, 'html.parser')
    summarytable = teamsoup.find('table', {'class': 'sortable stats_table'}, id=short_name)

    # Create arrays to hold info, will be in descending order from most current season to oldest season
    years_active_urls = []
    team_data = []

    # Go through each row of summary table
    for row in summarytable.findAll('tr'):

        # Get data
        cells = row.findAll('td')
        columns = row.findAll('th')
        if not len(cells) == 0 and not len(columns) == 0:
            for item in columns:
                years_active_urls.append(config.baseurl + item.find('a', href=True)['href'])
        team_data.append(config_tanya.scraper1_header.get_row_data(cells, columns))

    # Write team data to file
    dest = config.dest + short_name + '.xlsx'
    shutil.copyfile(config.template, dest)
    workbook = op.load_workbook(dest)
    sheet = workbook['overall_info']
    for i in range(len(team_data)):
        for j in range(len(team_data[i])):
            sheet.cell(row=i + 1, column=j + 1).value = team_data[i][j]
    workbook.save(dest)

    # Find team members for most current season
    seasonurl = years_active_urls[0]
    seasonpage = urllib.request.urlopen(seasonurl)
    seasonsoup = BeautifulSoup(seasonpage, 'html.parser')
    rostertable = seasonsoup.find('table', {'class': 'sortable stats_table'}, id='roster')

    # Create arrays to hold summarized info
    player_urls = []
    player_sumdata = []

    # Go through each row of roster table
    for row in rostertable.findAll('tr'):

        # Get data
        cells = row.findAll('td')
        columns = row.findAll('th')
        if not len(cells) == 0 and not len(columns) == 0:
            player_urls.append(config.baseurl + cells[0].find('a', href=True)['href'])
        player_sumdata.append(config_tanya.scraper1_header.get_row_data(cells, columns))

    # Write roster data to file
    sheet = workbook['current_roster']
    for i in range(len(player_sumdata)):
        for j in range(len(player_sumdata[i])):
            sheet.cell(row=i + 1, column=j + 1).value = player_sumdata[i][j]
    workbook.save(dest)
    workbook.close()

    # List all relevant metrics for a player
    player_metrics_titles = config.playermetrics
    player_metrics = []

    # Iterate through each player
    for i in range(len(player_urls)):

        currentplayer = []

        try:

            # Find detailed information for each player
            playerpage = urllib.request.urlopen(player_urls[i])
            playersoup = BeautifulSoup(playerpage, 'html.parser')

            # Get all placeholder texts
            all_comments = playersoup.find_all(string=lambda text: isinstance(text, Comment))
            table_comments = ['']
            table_titles = ['standard']
            for c in all_comments:
                if 'skaters_advanced' in c:
                    table_comments.append(c)
                    table_titles.append('possession')
                elif 'stats_misc_plus_nhl' in c:
                    table_comments.append(c)
                    table_titles.append('miscellaneous')
                elif 'stats_basic_plus_nhl_po' in c:
                    table_comments.append(c)
                    table_titles.append('playoffs')
                elif 'stats_basic_minus_other' in c:
                    table_comments.append(c)
                    table_titles.append('other')
                elif 'similarity_scores' in c:
                    table_comments.append(c)
                    table_titles.append('similarity')
                elif 'penalty_shots' in c:
                    table_comments.append(c)
                    table_titles.append('penalty')
                elif 'stats_goalie_situational' in c:
                    table_comments.append(c)
                    table_titles.append('goalie_advanced')
                elif 'hat_tricks' in c:
                    table_comments.append(c)
                    table_titles.append('hat_tricks')
                elif 'playoff_ot_goals' in c:
                    table_comments.append(c)
                    table_titles.append('ot_goals')

            # Get player metrics
            for title in player_metrics_titles:
                currenttitle = []
                for j in range(len(table_titles)):
                    if table_titles[j] == title:
                        currenttitle = config_tanya.scraper1_header.get_table_data(title, playersoup, table_comments[j])
                currentplayer.append(currenttitle)
            player_metrics.append(currentplayer)
            print(player_sumdata[i + 1][1] + ' complete')

        except AttributeError:

            # If a player has no historical data
            print(player_sumdata[i + 1][1] + " has no information available")
            player_metrics.append(currentplayer)

    for i in range(len(player_metrics)):

        # Write player data to file
        playerdest = dest = config.dest + 'players/' + player_sumdata[i + 1][1] + '.xlsx'
        shutil.copyfile(config.playertemplate, playerdest)
        workbook = op.load_workbook(dest)
        if len(player_metrics[i]) > 0:
            for metric, title in zip(player_metrics[i], player_metrics_titles):
                workbook.create_sheet(title)
                sheet = workbook[title]
                for j in range(len(metric)):
                    for k in range(len(metric[0])):
                        sheet.cell(row=j + 1, column=k + 1).value = metric[j][k]
            workbook.remove(workbook['temp'])
        workbook.save(playerdest)
        workbook.close()


def scrape2(url1, startingyear, endyear):

    print('Starting scraper 2 script')

    # Connect to SQL database
    my_db, my_cursor = connect_db('50_In_07')

    # Make tables for player game data if does not exist
    players, goalies = get_team_players(url1, '2019')
    print('Got player ids')
    for key in players:
        for player in players[key]:
            my_cursor.execute(f"""SHOW TABLES LIKE '{player}_games'""")
            results = my_cursor.fetchone()
            if results:
                continue
            else:
                if player not in goalies[key]:
                    my_cursor.execute(f"""CREATE TABLE {player}_games (
                    id VARCHAR(20) NOT NULL,
                    year YEAR NOT NULL,
                    player_team VARCHAR(7) NOT NULL,
                    opposing_team VARCHAR(7) NOT NULL,
                    goals INT NOT NULL,
                    assists INT NOT NULL,
                    points INT NOT NULL,
                    penalties_in_min INT NOT NULL,
                    time_on_ice TIME(0) NOT NULL)""")
                else:
                    my_cursor.execute(f"""CREATE TABLE {player}_games (
                    id VARCHAR(20) NOT NULL,
                    year YEAR NOT NULL,
                    player_team VARCHAR(7) NOT NULL,
                    opposing_team VARCHAR(7) NOT NULL,
                    goals_against INT NOT NULL,
                    shots_against INT NOT NULL,
                    saves INT NOT NULL,
                    penalties_in_min INT NOT NULL,
                    time_on_ice TIME(0) NOT NULL)""")
    print('SQL tables up to date')

    # Get urls for league summary for each year
    leagueurl = url1 + '/leagues/'
    leaguepage = urllib.request.urlopen(leagueurl)
    leaguesoup = BeautifulSoup(leaguepage, 'html.parser')
    leaguetable = leaguesoup.find('table', {'class': 'suppress_all sortable stats_table'}, id='league_index')
    yearurls = {}
    for row in leaguetable.findAll('tr'):
        columns = row.findAll('th')
        cells = row.findAll('td')
        if columns[0]['class'][0] == 'left' and len(cells) > 0 and cells[0].find(href=True):
            yearurls[columns[0].find(text=True)[0:4]] = cells[0].find(href=True)['href']
    print('Got urls for league summary pages')

    # Get urls for schedule/results for each year
    scheduleurls = {}
    for key in yearurls:
        if startingyear <= int(key) <= endyear:
            currenturl = url1 + yearurls[key]
            currentpage = urllib.request.urlopen(currenturl)
            currentsoup = BeautifulSoup(currentpage, 'html.parser')
            alllinks = currentsoup.findAll('a')
            for link in alllinks:
                if 'games' in link['href']:
                    scheduleurls[int(key)] = link['href']
                    break
        elif int(key) == 2009:
            break
    print('Got urls for schedule/results pages')

    # Get urls for each game
    gameurls = {}
    for key in scheduleurls:
        currenturl = url1 + scheduleurls[key]
        currentpage = urllib.request.urlopen(currenturl)
        currentsoup = BeautifulSoup(currentpage, 'html.parser')
        gametable = currentsoup.find('table', {'class': 'sortable stats_table'}, id='games')
        temp = []
        for row in gametable.findAll('tr'):
            columns = row.findAll('th')
            cells = row.findAll('td')
            if columns[0]['class'][0] == 'left' and columns[0].find(href=True):
                temp.append([columns[0].find(href=True)['href'],
                             cells[0]['csk'][0:3],
                             cells[2]['csk'][0:3]])
        playofftable = currentsoup.find('table', {'class': 'sortable stats_table'}, id='games_playoffs')
        for row in playofftable.findAll('tr'):
            columns = row.findAll('th')
            cells = row.findAll('td')
            if columns[0]['class'][0] == 'left' and columns[0].find(href=True):
                temp.append([columns[0].find(href=True)['href'],
                             cells[0]['csk'][0:3],
                             cells[2]['csk'][0:3]])
        gameurls[key] = temp
    print('Got urls for individual game pages')

    # Iterate through each game and add entry for each player
    for key in gameurls:
        for item in gameurls[key]:

            # Get game url
            currenturl = url1 + item[0]
            visitor = item[1]
            if visitor == 'PHX':
                visitor = 'ARI'
            elif visitor == 'ATL':
                visitor = 'WPG'
            home = item[2]
            if home == 'PHX':
                home = 'ARI'
            elif home == 'ATL':
                home = 'WPG'
            currentpage = urllib.request.urlopen(currenturl)
            currentsoup = BeautifulSoup(currentpage, 'html.parser')

            # Game data for each regular skater on visitor team
            visitortable = currentsoup.find('table', {'class': 'sortable stats_table'}, id=f'{visitor}_skaters')
            for line in visitortable.findAll('tr')[2:-1]:
                playername = line.findAll('td')[0].get('data-append-csv').replace('.', '') + '_games'
                try:
                    my_cursor.execute(f"""INSERT INTO {playername} VALUES (
                    '{item[0][11:-5]}',
                    {key},
                    '{item[1]}',
                    '{item[2]}',
                    {line.findAll('td')[1].find(text=True)},
                    {line.findAll('td')[2].find(text=True)},
                    {line.findAll('td')[3].find(text=True)},
                    {line.findAll('td')[5].find(text=True)},
                    '{line.findAll('td')[13].find(text=True)}:00')""")
                except Exception as e:
                    print(e)

            # Game data for each regular skater on home team
            hometable = currentsoup.find('table', {'class': 'sortable stats_table'}, id=f'{home}_skaters')
            for line in hometable.findAll('tr')[2:-1]:
                playername = line.findAll('td')[0].get('data-append-csv').replace('.', '') + '_games'
                try:
                    my_cursor.execute(f"""INSERT INTO {playername} VALUES (
                    '{item[0][11:-5]}',
                    {key},
                    '{item[2]}',
                    '{item[1]}',
                    {line.findAll('td')[1].find(text=True)},
                    {line.findAll('td')[2].find(text=True)},
                    {line.findAll('td')[3].find(text=True)},
                    {line.findAll('td')[5].find(text=True)},
                    '{line.findAll('td')[13].find(text=True)}:00')""")
                except Exception as e:
                    print(e)

            # Game data for visitor goalie(s)
            visitorgoalie = currentsoup.find('table', {'class': 'sortable stats_table'}, id=f'{visitor}_goalies')
            for line in visitorgoalie.findAll('tr')[2:]:
                visitorgoaliename = line.findAll('td')[0].get('data-append-csv').replace('.', '') + '_games'
                try:
                    my_cursor.execute(f"""INSERT INTO {visitorgoaliename} VALUES (
                    '{item[0][11:-5]}',
                    {key},
                    '{item[1]}',
                    '{item[2]}',
                    {line.findAll('td')[2].find(text=True)},
                    {line.findAll('td')[3].find(text=True)},
                    {line.findAll('td')[4].find(text=True)},
                    {line.findAll('td')[7].find(text=True)},
                    '{line.findAll('td')[8].find(text=True)}:00')""")
                except Exception as e:
                    print(e)

            # Game data for home goalie(s)
            homegoalie = currentsoup.find('table', {'class': 'sortable stats_table'}, id=f'{home}_goalies')
            for line in homegoalie.findAll('tr')[2:]:
                homegoaliename = line.findAll('td')[0].get('data-append-csv').replace('.', '') + '_games'
                try:
                    my_cursor.execute(f"""INSERT INTO {homegoaliename} VALUES (
                    '{item[0][11:-5]}',
                    {key},
                    '{item[2]}',
                    '{item[1]}',
                    {line.findAll('td')[2].find(text=True)},
                    {line.findAll('td')[3].find(text=True)},
                    {line.findAll('td')[4].find(text=True)},
                    {line.findAll('td')[7].find(text=True)},
                    '{line.findAll('td')[8].find(text=True)}:00')""")
                except Exception as e:
                    print(e)

            print(f'Completed game ' + item[0][11:-5])
        print(f'Completed year {key}')
    print('All game information added')

    # Disconnect from SQL database
    disconnect_db(my_db, my_cursor)


if __name__ == '__main__':

    # Get team names
    baseurl = 'https://www.hockey-reference.com'
    teams_short = import_shortnames(baseurl)
    teams_long = import_longnames(baseurl)

    # Run desired scraper
    n = int(input('Enter version of scraper desired: '))
    if n == 1:

        # Set config object
        baseurl = 'https://www.hockey-reference.com'
        date = '101219'
        config1 = ScraperConfig()
        config1.setconfig(baseurl, date)

        # Create team urls
        teams_url = []
        for team in teams_short:
            url = baseurl + '/teams/' + team + '/history.html'
            teams_url.append(url)

        # Get team info
        os.mkdir(config1.dest)
        os.mkdir(config1.dest + 'players/')
        for url, name in zip(teams_url, teams_short):
            scrape1(config1, url, name)

    elif n == 2:

        # Scrape data and insert into SQL database
        start = 2010
        end = 2014
        scrape2(baseurl, start, end)
