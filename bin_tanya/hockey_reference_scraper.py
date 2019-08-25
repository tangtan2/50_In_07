from bs4 import BeautifulSoup, Comment
import bin_tanya.header as header
import urllib.request
import shutil
import openpyxl as op


def scrape1(config, url, short_name):

    # Find overall info for team
    teampage = urllib.request.urlopen(url)
    teamsoup = BeautifulSoup(teampage, 'html.parser')
    summarytable = teamsoup.find('table', {'class': 'sortable stats_table'}, id=short_name)

    # Create arrays to hold info, will be in descending order from most current season to oldest season
    years_active_urls = []
    team_data = []

    # Go through each row of summary table
    for row in summarytable.findAll('tr'):

        # Get data
        cells = row.findAll('td')
        columns = row.findAll('th')
        if not len(cells) == 0 and not len(columns) == 0:
            for item in columns:
                years_active_urls.append(config.baseurl + item.find('a', href=True)['href'])
        team_data.append(header.get_row_data(cells, columns))

    # Write team data to file
    dest = config.dest + short_name + '.xlsx'
    shutil.copyfile(config.template, dest)
    workbook = op.load_workbook(dest)
    sheet = workbook['overall_info']
    for i in range(len(team_data)):
        for j in range(len(team_data[i])):
            sheet.cell(row=i + 1, column=j + 1).value = team_data[i][j]
    workbook.save(dest)

    # Find team members for most current season
    seasonurl = years_active_urls[0]
    seasonpage = urllib.request.urlopen(seasonurl)
    seasonsoup = BeautifulSoup(seasonpage, 'html.parser')
    rostertable = seasonsoup.find('table', {'class': 'sortable stats_table'}, id='roster')

    # Create arrays to hold summarized info
    player_urls = []
    player_sumdata = []

    # Go through each row of roster table
    for row in rostertable.findAll('tr'):

        # Get data
        cells = row.findAll('td')
        columns = row.findAll('th')
        if not len(cells) == 0 and not len(columns) == 0:
            player_urls.append(config.baseurl + cells[0].find('a', href=True)['href'])
        player_sumdata.append(header.get_row_data(cells, columns))

    # Write roster data to file
    sheet = workbook['current_roster']
    for i in range(len(player_sumdata)):
        for j in range(len(player_sumdata[i])):
            sheet.cell(row=i + 1, column=j + 1).value = player_sumdata[i][j]
    workbook.save(dest)
    workbook.close()

    # List all relevant metrics for a player
    player_metrics_titles = config.playermetrics
    player_metrics = []

    # Iterate through each player
    for i in range(len(player_urls)):

        currentplayer = []

        try:

            # Find detailed information for each player
            playerpage = urllib.request.urlopen(player_urls[i])
            playersoup = BeautifulSoup(playerpage, 'html.parser')

            # Get all placeholder texts
            all_comments = playersoup.find_all(string=lambda text: isinstance(text, Comment))
            table_comments = ['']
            table_titles = ['standard']
            for c in all_comments:
                if 'skaters_advanced' in c:
                    table_comments.append(c)
                    table_titles.append('possession')
                elif 'stats_misc_plus_nhl' in c:
                    table_comments.append(c)
                    table_titles.append('miscellaneous')
                elif 'stats_basic_plus_nhl_po' in c:
                    table_comments.append(c)
                    table_titles.append('playoffs')
                elif 'stats_basic_minus_other' in c:
                    table_comments.append(c)
                    table_titles.append('other')
                elif 'similarity_scores' in c:
                    table_comments.append(c)
                    table_titles.append('similarity')
                elif 'penalty_shots' in c:
                    table_comments.append(c)
                    table_titles.append('penalty')
                elif 'stats_goalie_situational' in c:
                    table_comments.append(c)
                    table_titles.append('goalie_advanced')
                elif 'hat_tricks' in c:
                    table_comments.append(c)
                    table_titles.append('hat_tricks')
                elif 'playoff_ot_goals' in c:
                    table_comments.append(c)
                    table_titles.append('ot_goals')

            # Get player metrics
            for title in player_metrics_titles:
                currenttitle = []
                for j in range(len(table_titles)):
                    if table_titles[j] == title:
                        currenttitle = header.get_table_data(title, playersoup, table_comments[j])
                currentplayer.append(currenttitle)
            player_metrics.append(currentplayer)
            print(player_sumdata[i + 1][1] + ' complete')

        except AttributeError:

            # If a player has no historical data
            print(player_sumdata[i + 1][1] + " has no information available")
            player_metrics.append(currentplayer)

    for i in range(len(player_metrics)):

        # Write player data to file
        playerdest = dest = config.dest + 'players/' + player_sumdata[i + 1][1] + '.xlsx'
        shutil.copyfile(config.playertemplate, playerdest)
        workbook = op.load_workbook(dest)
        if len(player_metrics[i]) > 0:
            for metric, title in zip(player_metrics[i], player_metrics_titles):
                workbook.create_sheet(title)
                sheet = workbook[title]
                for j in range(len(metric)):
                    for k in range(len(metric[0])):
                        sheet.cell(row=j + 1, column=k + 1).value = metric[j][k]
            workbook.remove(workbook['temp'])
        workbook.save(playerdest)
        workbook.close()
